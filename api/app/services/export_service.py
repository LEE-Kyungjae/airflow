"""
Export Service for data export functionality.

Provides CSV and Excel export capabilities with streaming support
for large datasets and asynchronous processing for batch exports.
"""

import csv
import io
import os
import uuid
import logging
from datetime import datetime, timedelta
from typing import AsyncIterator, Dict, List, Optional, Any, Generator
from bson import ObjectId

logger = logging.getLogger(__name__)

# Export file storage path
EXPORT_STORAGE_PATH = os.getenv('EXPORT_STORAGE_PATH', '/tmp/exports')
EXPORT_FILE_TTL_HOURS = int(os.getenv('EXPORT_FILE_TTL_HOURS', '24'))


class ExportService:
    """
    Service for exporting data to CSV and Excel formats.

    Supports both streaming exports for direct downloads and
    asynchronous exports for large datasets.

    Usage:
        export_service = ExportService(mongo_service)

        # Stream CSV
        async for chunk in export_service.stream_csv('crawl_results', query):
            yield chunk

        # Generate Excel
        excel_bytes = await export_service.generate_excel('crawl_results', query)
    """

    def __init__(self, mongo_service):
        """
        Initialize ExportService.

        Args:
            mongo_service: MongoService instance for database access
        """
        self.mongo = mongo_service
        self._ensure_storage_directory()

    def _ensure_storage_directory(self):
        """Ensure export storage directory exists."""
        if not os.path.exists(EXPORT_STORAGE_PATH):
            try:
                os.makedirs(EXPORT_STORAGE_PATH, exist_ok=True)
            except OSError as e:
                logger.warning(f"Could not create export storage directory: {e}")

    # ==================== CSV Export ====================

    def stream_csv(
        self,
        collection: str,
        query: Dict,
        fields: Optional[List[str]] = None,
        encoding: str = "utf-8-sig",
        batch_size: int = 1000,
        limit: int = 100000
    ) -> Generator[bytes, None, None]:
        """
        Generate CSV data as a stream.

        Yields CSV content in chunks for memory-efficient streaming downloads.
        Uses batch processing to handle large datasets.

        Args:
            collection: MongoDB collection name
            query: MongoDB query filter
            fields: List of fields to include (None for all)
            encoding: Character encoding (utf-8, utf-8-sig, euc-kr)
            batch_size: Number of records to fetch per batch
            limit: Maximum total records to export

        Yields:
            Bytes chunks of CSV content
        """
        try:
            # Get first document to determine headers
            first_doc = self.mongo.db[collection].find_one(query)
            if not first_doc:
                # Empty result - yield empty CSV with minimal header
                yield "No data found\n".encode(encoding)
                return

            # Determine fields to export
            if fields:
                headers = fields
            else:
                headers = self._get_document_headers(first_doc)

            # Create CSV writer with string buffer
            output = io.StringIO()
            writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

            # Write header row
            writer.writerow(headers)
            yield output.getvalue().encode(encoding)
            output.seek(0)
            output.truncate(0)

            # Stream data in batches
            skip = 0
            total_exported = 0

            while total_exported < limit:
                current_batch_size = min(batch_size, limit - total_exported)
                cursor = self.mongo.db[collection].find(query).skip(skip).limit(current_batch_size)
                documents = list(cursor)

                if not documents:
                    break

                for doc in documents:
                    row = self._document_to_row(doc, headers)
                    writer.writerow(row)

                yield output.getvalue().encode(encoding)
                output.seek(0)
                output.truncate(0)

                total_exported += len(documents)
                skip += len(documents)

                # Log progress for large exports
                if total_exported % 10000 == 0:
                    logger.info(f"CSV export progress: {total_exported} records")

            logger.info(f"CSV export completed: {total_exported} records from {collection}")

        except Exception as e:
            logger.error(f"CSV export error: {e}")
            raise

    async def stream_csv_async(
        self,
        collection: str,
        query: Dict,
        fields: Optional[List[str]] = None,
        encoding: str = "utf-8-sig",
        batch_size: int = 1000,
        limit: int = 100000
    ) -> AsyncIterator[bytes]:
        """
        Async version of CSV streaming for FastAPI StreamingResponse.

        Args:
            collection: MongoDB collection name
            query: MongoDB query filter
            fields: List of fields to include
            encoding: Character encoding
            batch_size: Records per batch
            limit: Maximum records

        Yields:
            Bytes chunks of CSV content
        """
        for chunk in self.stream_csv(collection, query, fields, encoding, batch_size, limit):
            yield chunk

    # ==================== Excel Export ====================

    def generate_excel(
        self,
        collection: str,
        query: Dict,
        fields: Optional[List[str]] = None,
        sheet_name: str = "Data",
        limit: int = 50000
    ) -> bytes:
        """
        Generate Excel file in memory.

        Creates an Excel workbook with the query results.
        Note: Excel generation loads all data into memory, so limit is important.

        Args:
            collection: MongoDB collection name
            query: MongoDB query filter
            fields: List of fields to include
            sheet_name: Name for the Excel worksheet
            limit: Maximum records (lower than CSV due to memory)

        Returns:
            Excel file content as bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        # Fetch data
        cursor = self.mongo.db[collection].find(query).limit(limit)
        documents = list(cursor)

        if not documents:
            # Return empty workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["No data found"])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        # Determine headers
        if fields:
            headers = fields
        else:
            headers = self._get_document_headers(documents[0])

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Style for header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, doc in enumerate(documents, 2):
            row = self._document_to_row(doc, headers)
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(header))

            # Sample first 100 rows for width calculation
            for row_idx in range(2, min(102, len(documents) + 2)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, min(len(str(cell_value)), 50))

            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel export completed: {len(documents)} records from {collection}")
        return output.getvalue()

    async def generate_excel_async(
        self,
        collection: str,
        query: Dict,
        fields: Optional[List[str]] = None,
        sheet_name: str = "Data",
        limit: int = 50000
    ) -> bytes:
        """
        Async wrapper for Excel generation.

        Args:
            collection: MongoDB collection name
            query: MongoDB query filter
            fields: List of fields to include
            sheet_name: Worksheet name
            limit: Maximum records

        Returns:
            Excel file content as bytes
        """
        return self.generate_excel(collection, query, fields, sheet_name, limit)

    # ==================== Async Export Jobs ====================

    def create_export_job(
        self,
        collection: str,
        format: str,
        query: Dict,
        fields: Optional[List[str]] = None,
        limit: int = 100000,
        encoding: str = "utf-8-sig"
    ) -> str:
        """
        Create an async export job record.

        Saves job metadata to database for background processing.

        Args:
            collection: Target collection
            format: Export format (csv, excel, json)
            query: MongoDB query filter
            fields: Fields to export
            limit: Maximum records
            encoding: Character encoding (for CSV)

        Returns:
            Job ID string
        """
        job_id = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"

        # Estimate record count
        estimated_count = min(
            self.mongo.db[collection].count_documents(query),
            limit
        )

        job_doc = {
            "_id": job_id,
            "collection": collection,
            "format": format,
            "query": query,
            "fields": fields,
            "limit": limit,
            "encoding": encoding,
            "status": "pending",
            "progress": 0.0,
            "records_processed": 0,
            "total_records": estimated_count,
            "file_path": None,
            "file_size": None,
            "error": None,
            "created_at": datetime.utcnow(),
            "started_at": None,
            "completed_at": None,
            "expires_at": None
        }

        self.mongo.db.export_jobs.insert_one(job_doc)
        logger.info(f"Created export job: {job_id} for {collection} ({estimated_count} records)")

        return job_id

    def get_export_job(self, job_id: str) -> Optional[Dict]:
        """
        Get export job status.

        Args:
            job_id: Job identifier

        Returns:
            Job document or None
        """
        doc = self.mongo.db.export_jobs.find_one({"_id": job_id})
        if doc:
            # Convert _id to job_id for response
            doc["job_id"] = doc.pop("_id")
        return doc

    def process_export_job(self, job_id: str) -> None:
        """
        Process an export job (runs in background).

        Executes the export and saves result to file system.
        Updates job status throughout processing.

        Args:
            job_id: Job identifier to process
        """
        job = self.mongo.db.export_jobs.find_one({"_id": job_id})
        if not job:
            logger.error(f"Export job not found: {job_id}")
            return

        try:
            # Update status to processing
            self.mongo.db.export_jobs.update_one(
                {"_id": job_id},
                {"$set": {
                    "status": "processing",
                    "started_at": datetime.utcnow()
                }}
            )

            collection = job["collection"]
            format = job["format"]
            query = job.get("query", {})
            fields = job.get("fields")
            limit = job.get("limit", 100000)
            encoding = job.get("encoding", "utf-8-sig")

            # Determine file extension
            ext = ".csv" if format == "csv" else ".xlsx" if format == "excel" else ".json"
            file_name = f"{job_id}{ext}"
            file_path = os.path.join(EXPORT_STORAGE_PATH, file_name)

            # Export based on format
            if format == "csv":
                self._process_csv_export(job_id, collection, query, fields, limit, encoding, file_path)
            elif format == "excel":
                self._process_excel_export(job_id, collection, query, fields, limit, file_path)
            elif format == "json":
                self._process_json_export(job_id, collection, query, fields, limit, file_path)
            else:
                raise ValueError(f"Unsupported format: {format}")

            # Get file size
            file_size = os.path.getsize(file_path)
            expires_at = datetime.utcnow() + timedelta(hours=EXPORT_FILE_TTL_HOURS)

            # Update job as completed
            self.mongo.db.export_jobs.update_one(
                {"_id": job_id},
                {"$set": {
                    "status": "completed",
                    "progress": 1.0,
                    "file_path": file_path,
                    "file_size": file_size,
                    "completed_at": datetime.utcnow(),
                    "expires_at": expires_at
                }}
            )

            logger.info(f"Export job completed: {job_id} ({file_size} bytes)")

        except Exception as e:
            logger.error(f"Export job failed: {job_id} - {e}")
            self.mongo.db.export_jobs.update_one(
                {"_id": job_id},
                {"$set": {
                    "status": "failed",
                    "error": str(e),
                    "completed_at": datetime.utcnow()
                }}
            )

    def _process_csv_export(
        self,
        job_id: str,
        collection: str,
        query: Dict,
        fields: Optional[List[str]],
        limit: int,
        encoding: str,
        file_path: str
    ) -> None:
        """Process CSV export to file."""
        with open(file_path, 'w', newline='', encoding=encoding) as f:
            writer = None
            headers = None
            total_processed = 0
            batch_size = 1000

            cursor = self.mongo.db[collection].find(query).limit(limit)

            for doc in cursor:
                if writer is None:
                    # Initialize writer with headers from first document
                    headers = fields if fields else self._get_document_headers(doc)
                    writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
                    writer.writerow(headers)

                row = self._document_to_row(doc, headers)
                writer.writerow(row)
                total_processed += 1

                # Update progress periodically
                if total_processed % batch_size == 0:
                    job = self.mongo.db.export_jobs.find_one({"_id": job_id})
                    total_records = job.get("total_records", limit)
                    progress = min(total_processed / total_records, 0.99) if total_records > 0 else 0

                    self.mongo.db.export_jobs.update_one(
                        {"_id": job_id},
                        {"$set": {
                            "progress": progress,
                            "records_processed": total_processed
                        }}
                    )

            # Final update
            self.mongo.db.export_jobs.update_one(
                {"_id": job_id},
                {"$set": {"records_processed": total_processed}}
            )

    def _process_excel_export(
        self,
        job_id: str,
        collection: str,
        query: Dict,
        fields: Optional[List[str]],
        limit: int,
        file_path: str
    ) -> None:
        """Process Excel export to file."""
        excel_bytes = self.generate_excel(collection, query, fields, "Data", limit)

        with open(file_path, 'wb') as f:
            f.write(excel_bytes)

        # Update records processed
        cursor = self.mongo.db[collection].find(query).limit(limit)
        count = len(list(cursor))
        self.mongo.db.export_jobs.update_one(
            {"_id": job_id},
            {"$set": {"records_processed": count}}
        )

    def _process_json_export(
        self,
        job_id: str,
        collection: str,
        query: Dict,
        fields: Optional[List[str]],
        limit: int,
        file_path: str
    ) -> None:
        """Process JSON export to file."""
        import json

        projection = {field: 1 for field in fields} if fields else None
        cursor = self.mongo.db[collection].find(query, projection).limit(limit)

        total_processed = 0
        batch_size = 1000

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write('[\n')
            first = True

            for doc in cursor:
                if not first:
                    f.write(',\n')
                first = False

                # Serialize document
                serialized = self._serialize_document(doc)
                f.write(json.dumps(serialized, ensure_ascii=False, default=str))
                total_processed += 1

                # Update progress
                if total_processed % batch_size == 0:
                    job = self.mongo.db.export_jobs.find_one({"_id": job_id})
                    total_records = job.get("total_records", limit)
                    progress = min(total_processed / total_records, 0.99) if total_records > 0 else 0

                    self.mongo.db.export_jobs.update_one(
                        {"_id": job_id},
                        {"$set": {
                            "progress": progress,
                            "records_processed": total_processed
                        }}
                    )

            f.write('\n]')

        self.mongo.db.export_jobs.update_one(
            {"_id": job_id},
            {"$set": {"records_processed": total_processed}}
        )

    def get_export_file_path(self, job_id: str) -> Optional[str]:
        """
        Get the file path for a completed export job.

        Args:
            job_id: Job identifier

        Returns:
            File path if job is completed and file exists, None otherwise
        """
        job = self.mongo.db.export_jobs.find_one({"_id": job_id})
        if not job:
            return None

        if job.get("status") != "completed":
            return None

        file_path = job.get("file_path")
        if file_path and os.path.exists(file_path):
            return file_path

        return None

    def cleanup_expired_exports(self) -> int:
        """
        Remove expired export files and job records.

        Returns:
            Number of files cleaned up
        """
        now = datetime.utcnow()
        cleaned = 0

        # Find expired jobs
        expired_jobs = self.mongo.db.export_jobs.find({
            "expires_at": {"$lt": now}
        })

        for job in expired_jobs:
            file_path = job.get("file_path")
            if file_path and os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    cleaned += 1
                except OSError as e:
                    logger.warning(f"Failed to delete export file: {file_path} - {e}")

        # Update job status to expired
        self.mongo.db.export_jobs.update_many(
            {"expires_at": {"$lt": now}, "status": "completed"},
            {"$set": {"status": "expired"}}
        )

        if cleaned > 0:
            logger.info(f"Cleaned up {cleaned} expired export files")

        return cleaned

    # ==================== Helper Methods ====================

    def _get_document_headers(self, doc: Dict) -> List[str]:
        """
        Extract field names from a document for CSV/Excel headers.

        Handles nested documents by flattening with dot notation.

        Args:
            doc: MongoDB document

        Returns:
            List of field names
        """
        headers = []
        for key, value in doc.items():
            if key == "_id":
                headers.append("id")
            elif isinstance(value, dict):
                # Flatten nested dict
                for nested_key in value.keys():
                    headers.append(f"{key}.{nested_key}")
            else:
                headers.append(key)
        return headers

    def _document_to_row(self, doc: Dict, headers: List[str]) -> List[Any]:
        """
        Convert a MongoDB document to a row for CSV/Excel.

        Args:
            doc: MongoDB document
            headers: List of header field names

        Returns:
            List of values in header order
        """
        row = []
        for header in headers:
            if header == "id":
                value = doc.get("_id")
            elif "." in header:
                # Handle nested field
                parts = header.split(".")
                value = doc
                for part in parts:
                    if isinstance(value, dict):
                        value = value.get(part)
                    else:
                        value = None
                        break
            else:
                value = doc.get(header)

            # Sanitize value for CSV/Excel
            row.append(self._sanitize_value(value))

        return row

    def _sanitize_value(self, value: Any) -> Any:
        """
        Sanitize a value for CSV/Excel export.

        Handles special types like ObjectId, datetime, lists, dicts.

        Args:
            value: Original value

        Returns:
            Sanitized value suitable for export
        """
        if value is None:
            return ""
        elif isinstance(value, ObjectId):
            return str(value)
        elif isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(value, (list, dict)):
            import json
            return json.dumps(value, ensure_ascii=False, default=str)
        elif isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        elif isinstance(value, str):
            # Escape potential CSV injection
            if value.startswith(('=', '+', '-', '@', '\t', '\r')):
                return "'" + value
            return value
        else:
            return value

    def _serialize_document(self, doc: Dict) -> Dict:
        """
        Serialize a MongoDB document for JSON export.

        Args:
            doc: MongoDB document

        Returns:
            JSON-serializable dictionary
        """
        result = {}
        for key, value in doc.items():
            if key == "_id":
                result["id"] = str(value)
            elif isinstance(value, ObjectId):
                result[key] = str(value)
            elif isinstance(value, datetime):
                result[key] = value.isoformat()
            elif isinstance(value, dict):
                result[key] = self._serialize_document(value)
            elif isinstance(value, list):
                result[key] = [
                    self._serialize_document(item) if isinstance(item, dict) else
                    str(item) if isinstance(item, ObjectId) else
                    item.isoformat() if isinstance(item, datetime) else
                    item
                    for item in value
                ]
            else:
                result[key] = value
        return result

    def _flatten_dict(self, d: Dict, parent_key: str = '', sep: str = '.') -> Dict:
        """
        Flatten a nested dictionary.

        Args:
            d: Dictionary to flatten
            parent_key: Parent key prefix
            sep: Separator for nested keys

        Returns:
            Flattened dictionary
        """
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)


# ============================================================
# Specialized Export Functions
# ============================================================

class CrawlResultExporter:
    """
    Specialized exporter for crawl results.

    Provides optimized queries and field selection for crawl result data.
    """

    def __init__(self, export_service: ExportService):
        self.export_service = export_service
        self.mongo = export_service.mongo

    def build_query(
        self,
        source_id: Optional[str] = None,
        status: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None
    ) -> Dict:
        """Build MongoDB query for crawl results."""
        query = {}

        if source_id:
            try:
                query["source_id"] = ObjectId(source_id)
            except Exception:
                query["source_id"] = source_id

        if status:
            query["status"] = status

        if date_from or date_to:
            query["executed_at"] = {}
            if date_from:
                query["executed_at"]["$gte"] = date_from
            if date_to:
                query["executed_at"]["$lte"] = date_to

        return query

    def get_export_fields(
        self,
        include_data: bool = False,
        include_metadata: bool = False
    ) -> List[str]:
        """Get list of fields to export based on options."""
        fields = [
            "id", "source_id", "status", "record_count",
            "execution_time_ms", "executed_at"
        ]

        if include_metadata:
            fields.extend(["crawler_id", "run_id", "error_code", "error_message"])

        if include_data:
            fields.append("data")

        return fields


class ReviewDataExporter:
    """
    Specialized exporter for review data.

    Provides optimized queries and field selection for data review records.
    """

    def __init__(self, export_service: ExportService):
        self.export_service = export_service
        self.mongo = export_service.mongo

    def build_query(
        self,
        source_id: Optional[str] = None,
        status: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None
    ) -> Dict:
        """Build MongoDB query for review data."""
        query = {}

        if source_id:
            try:
                query["source_id"] = ObjectId(source_id)
            except Exception:
                query["source_id"] = source_id

        if status:
            query["review_status"] = status

        if date_from or date_to:
            query["created_at"] = {}
            if date_from:
                query["created_at"]["$gte"] = date_from
            if date_to:
                query["created_at"]["$lte"] = date_to

        return query

    def get_export_fields(
        self,
        include_corrections: bool = True,
        include_confidence: bool = True
    ) -> List[str]:
        """Get list of fields to export based on options."""
        fields = [
            "id", "source_id", "crawl_result_id", "review_status",
            "original_data", "reviewed_at", "created_at"
        ]

        if include_corrections:
            fields.extend(["corrected_data", "corrections", "notes"])

        if include_confidence:
            fields.extend([
                "confidence_score", "ocr_confidence", "ai_confidence",
                "needs_number_review"
            ])

        return fields
